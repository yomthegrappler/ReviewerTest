import openpyxl
from Utility.constant import Constant

class ExcelExec:

    def __init__(self):
        self.constant = Constant()
        self.work_book = openpyxl.load_workbook(self.constant.Path_Excel)

    def get_sheet_name(self, sheet_name):
        sheet = self.work_book[sheet_name]
        return sheet

    def get_row_count(self, sheet_name):
        sheet = self.get_sheet_name(sheet_name)
        return sheet.max_row

    def get_col_count(self, sheet_name):
        sheet = self.get_sheet_name(sheet_name)
        return sheet.max_column

    def get_cell_data(self, sheet_name, row_count, col_count):
        sheet = self.get_sheet_name(sheet_name)
        return sheet.cell(row_count, col_count).value

    def get_object_value(self, sheet_name, object_name):
        row_count = self.get_row_count(sheet_name)
        for i in range(1, row_count + 1):
            if object_name == "":
                break
            elif str(object_name) == str(self.get_cell_data(sheet_name, i, self.constant.Page_ObjectName_Col)):
                object_value = str(self.get_cell_data(sheet_name, i, self.constant.Identifier))
                return object_value
